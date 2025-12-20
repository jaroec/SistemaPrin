# Backend/app/api/v1/export.py
from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date
from typing import Optional
import uuid
import os

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.db.base import get_db
from app.core.security import role_required
from app.db import models


router = APIRouter(prefix="/exports", tags=["📤 Exportaciones"])

def parse_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()

def get_financial_summary(db, start: date | None, end: date | None):
    sales = db.query(models.sale.Sale).filter(models.sale.Sale.status != "ANULADO")
    expenses = db.query(models.expense.Expense)

    if start:
        sales = sales.filter(func.date(models.sale.Sale.created_at) >= start)
        expenses = expenses.filter(func.date(models.expense.Expense.created_at) >= start)

    if end:
        sales = sales.filter(func.date(models.sale.Sale.created_at) <= end)
        expenses = expenses.filter(func.date(models.expense.Expense.created_at) <= end)

    income = sales.with_entities(func.sum(models.sale.Sale.total_usd)).scalar() or 0
    expense = expenses.with_entities(func.sum(models.expense.Expense.amount_usd)).scalar() or 0

    return {
        "income": round(income, 2),
        "expense": round(expense, 2),
        "balance": round(income - expense, 2),
        "sales_count": sales.count(),
        "expenses_count": expenses.count(),
    }


@router.get("/excel")
def export_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(role_required("ADMIN")),
):
    start = parse_date(start_date)
    end = parse_date(end_date)

    summary = get_financial_summary(db, start, end)

    file_path = f"/mnt/data/financial_{uuid.uuid4()}.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "Resumen"

    ws.append(["Concepto", "Valor (USD)"])
    rows = [
        ["Ingresos", summary["income"]],
        ["Egresos", summary["expense"]],
        ["Balance", summary["balance"]],
        ["Ventas", summary["sales_count"]],
        ["Gastos", summary["expenses_count"]],
    ]

    for row in rows:
        ws.append(row)

    header = PatternFill("solid", fgColor="4F81BD")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    wb.save(file_path)

    return FileResponse(
        file_path,
        filename="Reporte_Financiero.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@router.get("/pdf")
def export_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(role_required("ADMIN")),
):
    start = parse_date(start_date)
    end = parse_date(end_date)

    summary = get_financial_summary(db, start, end)

    file_path = f"/mnt/data/financial_{uuid.uuid4()}.pdf"
    doc = SimpleDocTemplate(file_path, pagesize=A4)

    styles = getSampleStyleSheet()
    elements = []

    logo_path = "/mnt/data/logo.png"
    if os.path.exists(logo_path):
        elements.append(Image(logo_path, width=120, height=50))

    elements.append(Paragraph("Reporte Financiero", styles["Title"]))
    elements.append(Paragraph(
        f"Período: {start or 'Inicio'} → {end or 'Hoy'}",
        styles["Normal"]
    ))

    table_data = [
        ["Concepto", "Monto (USD)"],
        ["Ingresos", summary["income"]],
        ["Egresos", summary["expense"]],
        ["Balance", summary["balance"]],
    ]

    table = Table(table_data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    elements.append(table)
    doc.build(elements)

    return FileResponse(
        file_path,
        filename="Reporte_Financiero.pdf",
        media_type="application/pdf",
    )
