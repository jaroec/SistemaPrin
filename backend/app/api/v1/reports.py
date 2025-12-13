# backend/app/api/v1/reports.py
from fastapi import APIRouter, Depends, Query 
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date
from typing import Optional
import uuid
import os
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from app.db.base import get_db
from app.core.security import role_required
from app.db import models

# ======================================================================================
# PROBLEMA SOLUCIONADO: SE QUITA EL /api/v1 (SE AGREGA EN main.py)
# ======================================================================================
router = APIRouter(prefix="/reports", tags=["Reports"])

# ======================================================================================
# FUNCIONES AUXILIARES COMPLETAS
# ======================================================================================

def get_financial_summary(db, start_date=None, end_date=None):
    sale_q = db.query(models.sale.Sale).filter(models.sale.Sale.status != "ANULADO")
    exp_q = db.query(models.expense.Expense)

    if start_date:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        sale_q = sale_q.filter(func.date(models.sale.Sale.created_at) >= start)
        exp_q = exp_q.filter(func.date(models.expense.Expense.created_at) >= start)

    if end_date:
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
        sale_q = sale_q.filter(func.date(models.sale.Sale.created_at) <= end)
        exp_q = exp_q.filter(func.date(models.expense.Expense.created_at) <= end)

    total_income = sale_q.with_entities(func.sum(models.sale.Sale.total_usd)).scalar() or 0
    total_expense = exp_q.with_entities(func.sum(models.expense.Expense.amount_usd)).scalar() or 0

    return {
        "total_income_usd": float(round(total_income, 2)),
        "total_expense_usd": float(round(total_expense, 2)),
        "balance_usd": float(round(total_income - total_expense, 2)),
        "count_sales": sale_q.count(),
        "count_expenses": exp_q.count(),
    }


def get_accounts_receivable(db, status="all"):
    today = date.today()
    q = db.query(models.accounts_receivable.AccountsReceivable)

    if status == "pending":
        q = q.filter(models.accounts_receivable.AccountsReceivable.status == "PENDIENTE")
    elif status == "paid":
        q = q.filter(models.accounts_receivable.AccountsReceivable.status == "PAGADO")
    elif status == "overdue":
        q = q.filter(
            models.accounts_receivable.AccountsReceivable.status == "PENDIENTE",
            models.accounts_receivable.AccountsReceivable.due_date < today
        )

    results = q.order_by(models.accounts_receivable.AccountsReceivable.due_date.asc()).all()

    return [
        {
            "id": r.id,
            "client_id": r.client_id,
            "client_name": r.client.name if r.client else None,
            "pending_amount_usd": float(r.pending_amount_usd),
            "due_date": r.due_date.isoformat(),
            "status": r.status,
        }
        for r in results
    ]


def get_accounts_payable(db, status="all"):
    today = date.today()
    q = db.query(models.accounts_payable.AccountsPayable)

    if status == "pending":
        q = q.filter(models.accounts_payable.AccountsPayable.status == "PENDIENTE")
    elif status == "paid":
        q = q.filter(models.accounts_payable.AccountsPayable.status == "PAGADO")
    elif status == "overdue":
        q = q.filter(
            models.accounts_payable.AccountsPayable.status == "PENDIENTE",
            models.accounts_payable.AccountsPayable.due_date < today
        )

    results = q.order_by(models.accounts_payable.AccountsPayable.due_date.asc()).all()

    return [
        {
            "id": r.id,
            "supplier_id": r.supplier_id,
            "supplier_name": r.supplier.name if r.supplier else None,
            "pending_amount_usd": float(r.pending_amount_usd),
            "due_date": r.due_date.isoformat(),
            "status": r.status,
        }
        for r in results
    ]


# ======================================================================================
# ENDPOINTS BASE (JSON)
# ======================================================================================

@router.get("/summary")
def report_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
):
    return get_financial_summary(db, start_date, end_date)


@router.get("/accounts_receivable")
def report_accounts_receivable(
    status: str = "all",
    db: Session = Depends(get_db)
):
    return get_accounts_receivable(db, status)


@router.get("/accounts_payable")
def report_accounts_payable(
    status: str = "all",
    db: Session = Depends(get_db)
):
    return get_accounts_payable(db, status)


# ======================================================================================
# EXPORTACIÓN EXCEL
# ======================================================================================

@router.get("/export/excel")
def export_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status_rcv: str = "all",
    status_pbl: str = "all",
    db: Session = Depends(get_db),
    current_user=Depends(role_required("ADMIN"))
):

    file_name = f"/mnt/data/report_{uuid.uuid4()}.xlsx"
    wb = Workbook()

    summary = get_financial_summary(db, start_date, end_date)

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Resumen"

    ws.append(["Concepto", "Valor (USD)"])
    rows = [
        ["Ingresos", summary["total_income_usd"]],
        ["Egresos", summary["total_expense_usd"]],
        ["Balance", summary["balance_usd"]],
        ["Total de Ventas", summary["count_sales"]],
        ["Total de Gastos", summary["count_expenses"]],
    ]

    for r in rows:
        ws.append(r)

    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    for col in range(1, 3):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # Sheet 2: Accounts Receivable
    rc_list = get_accounts_receivable(db, status_rcv)
    ws2 = wb.create_sheet("Cuentas por Cobrar")
    ws2.append(["ID", "Cliente", "Monto Pendiente", "Vencimiento", "Estado"])

    for r in rc_list:
        ws2.append([
            r["id"], r["client_name"], r["pending_amount_usd"], r["due_date"], r["status"]
        ])

    # Sheet 3: Accounts Payable
    pbl_list = get_accounts_payable(db, status_pbl)
    ws3 = wb.create_sheet("Cuentas por Pagar")
    ws3.append(["ID", "Proveedor", "Monto Pendiente", "Vencimiento", "Estado"])

    for p in pbl_list:
        ws3.append([
            p["id"], p["supplier_name"], p["pending_amount_usd"], p["due_date"], p["status"]
        ])

    wb.save(file_name)

    return FileResponse(
        file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="Reporte_Financiero.xlsx"
    )


# ======================================================================================
# EXPORTACIÓN PDF
# ======================================================================================

@router.get("/export/pdf")
def export_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(role_required("ADMIN"))
):
    file_name = f"/mnt/data/report_{uuid.uuid4()}.pdf"
    doc = SimpleDocTemplate(file_name, pagesize=A4)

    summary = get_financial_summary(db, start_date, end_date)
    receivable = get_accounts_receivable(db, "all")
    payable = get_accounts_payable(db, "all")

    elements = []
    styles = getSampleStyleSheet()

    logo_path = "/mnt/data/logo.png"
    if os.path.exists(logo_path):
        elements.append(Image(logo_path, width=120, height=60))

    elements.append(Paragraph("Reporte Financiero", styles["Title"]))
    elements.append(Paragraph(f"Período: {start_date or 'N/A'} a {end_date or 'N/A'}", styles["Normal"]))

    summary_table = [
        ["Concepto", "Valor (USD)"],
        ["Ingresos", summary["total_income_usd"]],
        ["Egresos", summary["total_expense_usd"]],
        ["Balance", summary["balance_usd"]],
    ]

    t = Table(summary_table)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightblue),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold")
    ]))
    elements.append(t)

    elements.append(Paragraph("Cuentas por Cobrar", styles["Heading2"]))
    data_rcv = [["ID", "Cliente", "Monto", "Vencimiento", "Estado"]]
    data_rcv += [[r["id"], r["client_name"], r["pending_amount_usd"], r["due_date"], r["status"]] for r in receivable]
    elements.append(Table(data_rcv))

    elements.append(Paragraph("Cuentas por Pagar", styles["Heading2"]))
    data_pbl = [["ID", "Proveedor", "Monto", "Vencimiento", "Estado"]]
    data_pbl += [[p["id"], p["supplier_name"], p["pending_amount_usd"], p["due_date"], p["status"]] for p in payable]
    elements.append(Table(data_pbl))

    doc.build(elements)

    return FileResponse(
        file_name,
        media_type="application/pdf",
        filename="Reporte_Financiero.pdf"
    )


# ======================================================================================
# CANCELACIÓN DE VENTA
# ======================================================================================

@router.post("/pos/sales/{sale_id}/cancel")
def cancel_sale(
    sale_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(role_required("ADMIN"))
):
    sale = db.query(models.sale.Sale).filter(models.sale.Sale.id == sale_id).first()

    if not sale:
        return {"error": "Venta no encontrada"}

    if sale.status == "ANULADO":
        return {"message": "La venta ya está anulada"}

    sale.status = "ANULADO"
    db.commit()

    return {"message": "Venta anulada correctamente", "sale_id": sale_id}
